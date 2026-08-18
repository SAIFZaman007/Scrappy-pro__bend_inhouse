"""Export a finished job to CSV or XLSX.

The column list and their order are fixed by the spec and must not drift:

    id, name, brand, category, sub, price, oldPrice, stock, rating,
    reviews, badge, image, images, specs, desc

``id`` is a stable, human-readable running number per export (NEW_001, NEW_002...).
``images`` and ``specs`` are flattened to a single cell each with a " | " separator,
which survives a round trip through Excel and Google Sheets without mangling URLs.

Rows are streamed from Postgres in chunks so a 50,000 product export never loads
the whole result set into memory.
"""

from __future__ import annotations

import csv
import uuid
from collections.abc import AsyncIterator, Iterable
from datetime import UTC, datetime
from pathlib import Path
from typing import Any

from openpyxl import Workbook
from openpyxl.cell import WriteOnlyCell
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from sqlalchemy import select
from sqlalchemy.ext.asyncio import AsyncSession

from app.core.config import settings
from app.core.logging import get_logger
from app.models.entities import ExportFile, Product, ScrapeJob, Site

log = get_logger(__name__)

COLUMNS: list[str] = [
    "id",
    "name",
    "brand",
    "category",
    "sub",
    "price",
    "oldPrice",
    "stock",
    "rating",
    "reviews",
    "badge",
    "image",
    "images",
    "specs",
    "desc",
]

MULTI_VALUE_SEPARATOR = " | "
CHUNK_SIZE = 500
# Excel refuses cells longer than this.
EXCEL_CELL_LIMIT = 32_000


def _join(values: Iterable[str]) -> str:
    return MULTI_VALUE_SEPARATOR.join(v for v in values if v)


def _format_specs(specs: dict[str, Any] | None) -> str:
    if not specs:
        return ""
    return _join(f"{key}: {value}" for key, value in specs.items())


def _row(product: Product, row_id: str) -> list[Any]:
    return [
        row_id,
        product.name or "",
        product.brand or "",
        product.category_name or "",
        product.subcategory_name or "",
        float(product.price) if product.price is not None else "",
        float(product.old_price) if product.old_price is not None else "",
        product.stock or "",
        product.rating if product.rating is not None else "",
        product.reviews if product.reviews is not None else "",
        product.badge or "",
        product.image or "",
        _join(product.images or []),
        _format_specs(product.specs),
        (product.description or "").strip(),
    ]


async def _iter_rows(
    db: AsyncSession, job_id: uuid.UUID, prefix: str
) -> AsyncIterator[list[Any]]:
    """Yield export rows in stable order, in chunks."""
    offset = 0
    counter = 0
    while True:
        products = (
            (
                await db.execute(
                    select(Product)
                    .where(Product.job_id == job_id)
                    .order_by(Product.sequence, Product.id)
                    .offset(offset)
                    .limit(CHUNK_SIZE)
                )
            )
            .scalars()
            .all()
        )
        if not products:
            return
        for product in products:
            counter += 1
            yield _row(product, f"{prefix}_{counter:03d}")
        offset += CHUNK_SIZE


def _export_dir() -> Path:
    directory = Path(settings.EXPORT_DIR)
    directory.mkdir(parents=True, exist_ok=True)
    return directory


def _filename(site_key: str, job_id: uuid.UUID, fmt: str) -> str:
    stamp = datetime.now(UTC).strftime("%Y%m%d-%H%M")
    return f"{site_key}-{stamp}-{str(job_id)[:8]}.{fmt}"


async def build_export(db: AsyncSession, job_id: uuid.UUID, fmt: str = "csv") -> ExportFile:
    fmt = fmt.lower()
    if fmt not in ("csv", "xlsx"):
        raise ValueError("Format must be csv or xlsx")

    job = (await db.execute(select(ScrapeJob).where(ScrapeJob.id == job_id))).scalar_one()
    site = (await db.execute(select(Site).where(Site.id == job.site_id))).scalar_one()
    prefix = str((job.options or {}).get("id_prefix") or "NEW").upper()

    path = _export_dir() / _filename(site.key, job_id, fmt)
    rows_written = 0

    if fmt == "csv":
        # utf-8-sig so Excel on Windows opens Bengali text correctly.
        with path.open("w", newline="", encoding="utf-8-sig") as handle:
            writer = csv.writer(handle, quoting=csv.QUOTE_MINIMAL)
            writer.writerow(COLUMNS)
            async for row in _iter_rows(db, job_id, prefix):
                writer.writerow(row)
                rows_written += 1
    else:
        workbook = Workbook(write_only=True)
        sheet = workbook.create_sheet("Products")
        _size_columns(sheet)
        sheet.append(_header_cells(sheet))
        async for row in _iter_rows(db, job_id, prefix):
            sheet.append(
                [
                    (v[:EXCEL_CELL_LIMIT] if isinstance(v, str) else v)
                    for v in row
                ]
            )
            rows_written += 1
        workbook.save(path)

    record = ExportFile(
        job_id=job_id,
        fmt=fmt,
        filename=path.name,
        path=str(path),
        row_count=rows_written,
        size_bytes=path.stat().st_size,
        created_at=datetime.now(UTC),
    )
    db.add(record)
    await db.commit()
    await db.refresh(record)
    log.info("export.built", job_id=str(job_id), fmt=fmt, rows=rows_written)
    return record


COLUMN_WIDTHS = {
        "id": 12, "name": 52, "brand": 16, "category": 18, "sub": 18,
        "price": 12, "oldPrice": 12, "stock": 14, "rating": 9, "reviews": 10,
    "badge": 12, "image": 40, "images": 40, "specs": 60, "desc": 60,
}


def _size_columns(sheet: Any) -> None:
    """Widths and the frozen header must be set before any row is written."""
    for index, column in enumerate(COLUMNS, start=1):
        sheet.column_dimensions[get_column_letter(index)].width = COLUMN_WIDTHS.get(column, 18)
    sheet.freeze_panes = "A2"


def _header_cells(sheet: Any) -> list[Any]:
    """A write-only sheet styles the header by appending pre-styled cells."""
    fill = PatternFill("solid", fgColor="0E1726")
    font = Font(color="FFFFFF", bold=True)
    cells = []
    for column in COLUMNS:
        cell = WriteOnlyCell(sheet, value=column)
        cell.fill = fill
        cell.font = font
        cell.alignment = Alignment(vertical="center")
        cells.append(cell)
    return cells


def purge_expired_exports() -> int:
    """Delete export files older than the retention window. Called by the worker cron."""
    directory = Path(settings.EXPORT_DIR)
    if not directory.exists():
        return 0
    cutoff = datetime.now(UTC).timestamp() - settings.EXPORT_RETENTION_HOURS * 3600
    removed = 0
    for file in directory.iterdir():
        if file.is_file() and file.stat().st_mtime < cutoff:
            file.unlink(missing_ok=True)
            removed += 1
    return removed
